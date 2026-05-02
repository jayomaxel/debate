"""
报告生成服务
负责生成辩论报告、导出PDF/Excel
"""
import asyncio
import json

from logging_config import get_logger
from typing import List, Dict, Optional
from sqlalchemy.orm import Session
from sqlalchemy import select
from datetime import datetime
from io import BytesIO
from pathlib import Path

from models.debate import Debate, DebateParticipation
from models.speech import Speech
from models.score import Score
from models.user import User
from services.scoring_service import ScoringService
from services.config_service import ConfigService
from services.coze_client import CozeClient
from config import settings

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, XPreformatted
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from agents.judge_agent import JudgeAgent
from utils.markdown_to_pdf import markdown_to_pdf

logger = get_logger(__name__)


class Report:
    """报告数据结构"""
    
    def __init__(
        self,
        debate_id: str,
        topic: str,
        start_time: datetime,
        end_time: datetime,
        duration: int,
        participants: List[Dict],
        speeches: List[Dict],
        statistics: Dict,
        winner: str
    ):
        self.debate_id = debate_id
        self.topic = topic
        self.start_time = start_time
        self.end_time = end_time
        self.duration = duration
        self.participants = participants
        self.speeches = speeches
        self.statistics = statistics
        self.winner = winner
    
    def to_dict(self) -> dict:
        """转换为字典"""
        return {
            "debate_id": self.debate_id,
            "topic": self.topic,
            "start_time": self.start_time.isoformat() if self.start_time else None,
            "end_time": self.end_time.isoformat() if self.end_time else None,
            "duration": self.duration,
            "participants": self.participants,
            "speeches": self.speeches,
            "statistics": self.statistics,
            "winner": self.winner
        }


class ReportGenerator:
    """报告生成器"""
    
    @staticmethod
    def generate_student_report(
        db: Session,
        debate_id: str,
        student_id: str
    ) -> Optional[Report]:
        """
        生成学生报告
        
        Args:
            db: 数据库会话
            debate_id: 辩论ID
            student_id: 学生ID
            
        Returns:
            报告对象
        """
        try:
            # 获取辩论信息
            debate = db.execute(
                select(Debate).where(Debate.id == debate_id)
            ).scalar_one_or_none()
            
            if not debate:
                logger.error(f"辩论不存在: {debate_id}")
                return None
            
            viewer = db.execute(select(User).where(User.id == student_id)).scalar_one_or_none()
            if not viewer:
                logger.error(f"用户不存在: user_id={student_id}")
                return None

            if viewer.user_type == "student":
                participation = db.execute(
                    select(DebateParticipation).where(
                        DebateParticipation.debate_id == debate_id,
                        DebateParticipation.user_id == student_id
                    )
                ).scalar_one_or_none()

                if not participation:
                    logger.error(f"学生未参与该辩论: student_id={student_id}")
                    return None

            all_participations = (
                db.execute(
                    select(DebateParticipation).where(DebateParticipation.debate_id == debate_id)
                )
                .scalars()
                .all()
            )

            participation_by_user_id = {str(p.user_id): p for p in all_participations if p.user_id}
            role_to_stance = {str(p.role): str(p.stance) for p in all_participations if p.role and p.stance}

            user_ids = [p.user_id for p in all_participations if p.user_id]
            users: List[User] = []
            if user_ids:
                users = db.execute(select(User).where(User.id.in_(user_ids))).scalars().all()
            user_by_id = {str(u.id): u for u in users}

            speeches: List[Speech] = (
                db.execute(
                    select(Speech)
                    .where(Speech.debate_id == debate_id)
                    .where(Speech.is_valid_for_scoring.is_(True))
                    .order_by(Speech.timestamp)
                )
                .scalars()
                .all()
            )
            speeches = [
                speech
                for speech in speeches
                if str(getattr(speech, "content", "") or "").strip()
            ]

            score_rows = (
                db.execute(
                    select(Score, Speech)
                    .join(Speech, Score.speech_id == Speech.id)
                    .where(Speech.debate_id == debate_id)
                    .where(Speech.is_valid_for_scoring.is_(True))
                )
                .all()
            )
            score_by_speech_id: Dict[str, Score] = {
                str(speech.id): score for (score, speech) in score_rows if speech and score
            }

            def _ai_role_to_name(speaker_role: str) -> str:
                role = str(speaker_role)
                if role.startswith("ai_"):
                    parts = role.split("_", 1)
                    if len(parts) == 2 and parts[1].isdigit():
                        return f"AI辩手{parts[1]}"
                return "AI"

            def _ai_role_to_mapped_debater_role(speaker_role: str) -> Optional[str]:
                role = str(speaker_role)
                if role.startswith("ai_"):
                    parts = role.split("_", 1)
                    if len(parts) == 2 and parts[1].isdigit():
                        return f"debater_{parts[1]}"
                return None

            human_speeches_by_user: Dict[str, List[Speech]] = {}
            ai_speeches_by_role: Dict[str, List[Speech]] = {}

            for sp in speeches:
                if str(sp.speaker_type) == "human" and sp.speaker_id:
                    key = str(sp.speaker_id)
                    human_speeches_by_user.setdefault(key, []).append(sp)
                elif str(sp.speaker_type) == "ai":
                    key = str(sp.speaker_role)
                    ai_speeches_by_role.setdefault(key, []).append(sp)

            participants: List[Dict] = []

            def _compute_final_score(speech_list: List[Speech]) -> Dict:
                speech_count = len(speech_list)
                total_duration = sum(max(0, int(s.duration or 0)) for s in speech_list)
                scored = [score_by_speech_id.get(str(s.id)) for s in speech_list]
                scored = [s for s in scored if s is not None]
                if not scored:
                    return {
                        "logic_score": 0.0,
                        "argument_score": 0.0,
                        "response_score": 0.0,
                        "persuasion_score": 0.0,
                        "teamwork_score": 0.0,
                        "overall_score": 0.0,
                        "speech_count": speech_count,
                        "total_duration": total_duration,
                    }
                logic_avg = sum(float(s.logic_score) for s in scored) / len(scored)
                argument_avg = sum(float(s.argument_score) for s in scored) / len(scored)
                response_avg = sum(float(s.response_score) for s in scored) / len(scored)
                persuasion_avg = sum(float(s.persuasion_score) for s in scored) / len(scored)
                teamwork_avg = sum(float(s.teamwork_score) for s in scored) / len(scored)
                overall_avg = sum(float(s.overall_score) for s in scored) / len(scored)
                return {
                    "logic_score": round(logic_avg, 2),
                    "argument_score": round(argument_avg, 2),
                    "response_score": round(response_avg, 2),
                    "persuasion_score": round(persuasion_avg, 2),
                    "teamwork_score": round(teamwork_avg, 2),
                    "overall_score": round(overall_avg, 2),
                    "speech_count": speech_count,
                    "total_duration": total_duration,
                }

            for user_id_str, speech_list_for_user in human_speeches_by_user.items():
                user = user_by_id.get(user_id_str)
                participation_for_user = participation_by_user_id.get(user_id_str)
                role = (
                    str(participation_for_user.role)
                    if participation_for_user
                    else str(speech_list_for_user[0].speaker_role)
                )
                stance = (
                    str(participation_for_user.stance) if participation_for_user else None
                )
                participants.append(
                    {
                        "user_id": user_id_str,
                        "name": user.name if user else "未知用户",
                        "role": role,
                        "stance": stance,
                        "is_ai": False,
                        "final_score": _compute_final_score(speech_list_for_user),
                    }
                )

            for ai_role, speech_list_for_ai in ai_speeches_by_role.items():
                mapped_role = _ai_role_to_mapped_debater_role(ai_role)
                participants.append(
                    {
                        "user_id": f"ai:{ai_role}",
                        "name": _ai_role_to_name(ai_role),
                        "role": ai_role,
                        "stance": "negative",
                        "is_ai": True,
                        "final_score": _compute_final_score(speech_list_for_ai),
                    }
                )

            participant_name_by_human_id = {
                p["user_id"]: p["name"] for p in participants if not p.get("is_ai")
            }
            ai_name_by_role = {
                p["role"]: p["name"] for p in participants if p.get("is_ai")
            }

            speech_list: List[Dict] = []
            for sp in speeches:
                speaker_type = str(sp.speaker_type)
                speaker_role = str(sp.speaker_role)
                speaker_name = None
                stance = None
                role = None

                if speaker_type == "human" and sp.speaker_id:
                    speaker_name = participant_name_by_human_id.get(str(sp.speaker_id)) or (
                        user_by_id.get(str(sp.speaker_id)).name if user_by_id.get(str(sp.speaker_id)) else "未知用户"
                    )
                    p = participation_by_user_id.get(str(sp.speaker_id))
                    if p:
                        stance = str(p.stance)
                        role = str(p.role)
                    else:
                        role = speaker_role
                else:
                    speaker_name = ai_name_by_role.get(speaker_role) or _ai_role_to_name(speaker_role)
                    mapped_role = _ai_role_to_mapped_debater_role(speaker_role)
                    stance = role_to_stance.get(mapped_role) if mapped_role else None
                    role = speaker_role

                speech_score = score_by_speech_id.get(str(sp.id))

                speech_list.append(
                    {
                        "id": str(sp.id),
                        "speaker_type": speaker_type,
                        "speaker_role": speaker_role,
                        "speaker_name": speaker_name,
                        "stance": stance,
                        "role": role,
                        "phase": sp.phase,
                        "content": sp.content,
                        "duration": int(sp.duration or 0),
                        "timestamp": sp.timestamp.isoformat(),
                        "score": {
                            "logic_score": speech_score.logic_score,
                            "argument_score": speech_score.argument_score,
                            "response_score": speech_score.response_score,
                            "persuasion_score": speech_score.persuasion_score,
                            "teamwork_score": speech_score.teamwork_score,
                            "overall_score": speech_score.overall_score,
                            "feedback": speech_score.feedback or "",
                        }
                        if speech_score
                        else None,
                    }
                )
            
            # 获取辩论统计
            statistics = ScoringService.get_debate_statistics(db, debate_id)
            
            # 如果有全局评分报告，覆盖统计数据
            if debate.report:
                report_data = debate.report
                statistics["winner"] = report_data.get("winner", statistics.get("winner"))
                statistics["winning_reason"] = report_data.get("winning_reason")
                statistics["global_scores"] = report_data.get("scores")
                statistics["overall_comment"] = report_data.get("overall_comment")
                statistics["suggestions"] = report_data.get("suggestions")
            
            # 计算持续时间
            duration = 0
            if debate.start_time and debate.end_time:
                duration = int((debate.end_time - debate.start_time).total_seconds() / 60)
            
            # 创建报告
            report = Report(
                debate_id=str(debate.id),
                topic=debate.topic,
                start_time=debate.start_time,
                end_time=debate.end_time,
                duration=duration,
                participants=participants,
                speeches=speech_list,
                statistics=statistics,
                winner=statistics.get("winner", "unknown")
            )
            
            logger.info(f"学生报告生成成功: debate_id={debate_id}, student_id={student_id}")
            
            return report
        
        except Exception as e:
            logger.error(f"生成学生报告失败: {e}", exc_info=True)
            return None
    
    @staticmethod
    def generate_class_report(
        db: Session,
        class_id: str
    ) -> Optional[Dict]:
        """
        生成班级报告
        
        Args:
            db: 数据库会话
            class_id: 班级ID
            
        Returns:
            班级报告数据
        """
        try:
            # 获取班级的所有辩论
            debates = db.execute(
                select(Debate).where(
                    Debate.class_id == class_id,
                    Debate.status == "completed"
                )
            ).scalars().all()
            
            # 获取班级的所有学生
            students = db.execute(
                select(User).where(
                    User.class_id == class_id,
                    User.user_type == "student"
                )
            ).scalars().all()
            
            # 统计每个学生的数据
            student_stats = []
            
            for student in students:
                # 获取学生参与的辩论
                participations = db.execute(
                    select(DebateParticipation).where(
                        DebateParticipation.user_id == student.id
                    )
                ).scalars().all()
                
                # 计算平均分
                total_score = 0
                score_count = 0
                
                for participation in participations:
                    final_score = ScoringService.calculate_final_score(
                        db, str(participation.id)
                    )
                    total_score += final_score["overall_score"]
                    score_count += 1
                
                avg_score = total_score / score_count if score_count > 0 else 0
                
                # 统计发言次数
                speech_count = db.execute(
                    select(Speech).where(Speech.speaker_id == student.id)
                ).scalars().all()
                
                student_stats.append({
                    "user_id": str(student.id),
                    "name": student.name,
                    "student_id": student.student_id,
                    "participation_count": len(participations),
                    "speech_count": len(speech_count),
                    "avg_score": round(avg_score, 2)
                })
            
            # 班级整体统计
            class_stats = {
                "total_students": len(students),
                "total_debates": len(debates),
                "completed_debates": len([d for d in debates if d.status == "completed"]),
                "avg_participation_rate": 0,
                "avg_class_score": 0
            }
            
            if students:
                total_participations = sum(s["participation_count"] for s in student_stats)
                class_stats["avg_participation_rate"] = round(
                    total_participations / len(students),
                    2
                )
                
                total_scores = sum(s["avg_score"] for s in student_stats if s["avg_score"] > 0)
                students_with_scores = len([s for s in student_stats if s["avg_score"] > 0])
                
                if students_with_scores > 0:
                    class_stats["avg_class_score"] = round(
                        total_scores / students_with_scores,
                        2
                    )
            
            logger.info(f"班级报告生成成功: class_id={class_id}")
            
            return {
                "class_id": class_id,
                "students": student_stats,
                "statistics": class_stats,
                "generated_at": datetime.utcnow().isoformat()
            }
        
        except Exception as e:
            logger.error(f"生成班级报告失败: {e}", exc_info=True)
            return None
    
    @staticmethod
    def export_to_pdf(report: Report) -> Optional[bytes]:
        raise RuntimeError("export_to_pdf requires db; use export_to_pdf_async")

    @staticmethod
    def _get_report_pdf_cache_path(debate_id: str) -> Path:
        base = Path(settings.UPLOAD_DIR)
        return base / "reports" / f"debate_report_{debate_id}.pdf"

    @staticmethod
    def _build_fallback_markdown(report: Report) -> str:
        participants_md = "\n".join(
            [
                f"- {p.get('name')}（{p.get('role')}）得分：{(p.get('final_score') or {}).get('overall_score', 0)}"
                for p in (report.participants or [])
            ]
        )
        speeches_md_lines: List[str] = []
        for i, s in enumerate(report.speeches or []):
            score = (s.get("score") or {}).get("overall_score") if s.get("score") else None
            score_txt = f"{score}" if score is not None else "N/A"
            speaker_name = s.get("speaker_name") or s.get("speaker_role") or "未知"
            phase = s.get("phase") or ""
            speeches_md_lines.append(f"{i+1}. **[{phase}] {speaker_name}**（得分：{score_txt}）\n\n{s.get('content') or ''}\n")
        speeches_md = "\n".join(speeches_md_lines)

        return (
            f"# 辩论报告\n\n"
            f"## 基本信息\n\n"
            f"- 辩题：{report.topic}\n"
            f"- 开始时间：{report.start_time}\n"
            f"- 结束时间：{report.end_time}\n"
            f"- 持续时间：{report.duration}分钟\n"
            f"- 获胜方：{report.winner}\n\n"
            f"## 参与者\n\n"
            f"{participants_md}\n\n"
            f"## 发言记录\n\n"
            f"{speeches_md}\n"
        )

    @staticmethod
    def _build_coze_prompt(report: Report) -> str:
        stats = report.statistics or {}
        winner = stats.get("winner") or report.winner
        topic = report.topic
        duration = report.duration
        start = report.start_time
        end = report.end_time
        return (
            "你是辩论裁判AI，请基于“对话消息记录”生成一份可直接用于教学复盘的《辩论完整报告》。\n"
            "输出要求：\n"
            "1) 只输出 Markdown 正文，不要代码块包裹，不要额外解释。\n"
            "2) 报告需包含：概览（辩题/时间/胜负）、双方核心论点提炼、关键回合与转折点、逻辑建构力分析、AI核心知识运用评估、批判性思维表现、语言表达力点评、AI伦理与科技素养体现、逐个辩手的优缺点与改进建议、全局建议。\n"
            "3) 在评价每位辩手时，请引用其使用的AI课程术语（如情感计算、NLP、AIGC、AI伦理等），并对术语使用的准确性和深度给出具体评价。\n"
            "4) 若消息里带有每段发言的评分与理由，请在“详细分析”里按发言序号引用并整合。\n"
            "5) 语言：中文，行文客观、可操作。\n\n"
            f"辩题：{topic}\n"
            f"开始时间：{start}\n"
            f"结束时间：{end}\n"
            f"时长：{duration}分钟\n"
            f"当前统计胜者：{winner}\n"
        )

    @staticmethod
    async def _generate_markdown_via_coze(db: Session,debate_id:str, message_str: str) -> str:
        try:
            config_service = ConfigService(db)
            coze_config = await config_service.get_coze_config()
            bot_id = (getattr(coze_config, "judge_bot_id", "") or "").strip()
            api_token = (getattr(coze_config, "api_token", "") or "").strip()
            params = getattr(coze_config, "parameters", None) or {}
            base_url = (params.get("base_url") or settings.COZE_BASE_URL or "").strip()
            if not (bot_id and api_token and base_url):
                return ""

            coze = CozeClient(api_token=api_token, base_url=base_url)
            raw_messages: List[Dict] = [{"role": "user", "content": message_str}]


            raw_messages.append({"role": "user", "content": "请开始生成《辩论完整报告》Markdown。"})
            chat_params = coze.build_chat_coze_params(
                bot_id=bot_id,
                user_id=f"report:{debate_id}",
                raw_messages=raw_messages,
            )
            text = await asyncio.to_thread(
                coze.chat_coze,
                bot_id=chat_params["bot_id"],
                user_id=chat_params["user_id"],
                messages=chat_params["messages"],
                extra=chat_params.get("extra"),
            )
            return (text or "").strip()
        except Exception as e:
            logger.error(f"生成Coze Markdown报告失败: {e}", exc_info=True)
            return ""

    @staticmethod
    def _markdown_to_pdf_bytes(markdown_text: str, *, title: str) -> bytes:
        """
        使用WeasyPrint将Markdown转换为PDF（已弃用，保留用于兼容）
        建议使用 _markdown_to_pdf_bytes_weasyprint
        """
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "TitleCN",
            parent=styles["Title"],
            fontName="STSong-Light",
            fontSize=18,
            leading=22,
            spaceAfter=8 * mm,
        )
        body_style = ParagraphStyle(
            "BodyCN",
            parent=styles["BodyText"],
            fontName="STSong-Light",
            fontSize=11,
            leading=15,
            wordWrap="CJK",
        )
        mono_style = ParagraphStyle(
            "MonoCN",
            parent=styles["Code"],
            fontName="STSong-Light",
            fontSize=10.5,
            leading=14,
            wordWrap="CJK",
        )

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=18 * mm,
            rightMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
            title=title,
        )
        story = [Paragraph(title, title_style), Spacer(1, 4 * mm)]
        text = (markdown_text or "").strip()
        if not text:
            text = "报告内容为空"
        story.append(XPreformatted(text, mono_style))
        story.append(Spacer(1, 2 * mm))
        doc.build(story)
        return buffer.getvalue()
    
    @staticmethod
    async def _markdown_to_pdf_bytes_weasyprint(
        markdown_text: str,
        title: str,
        debate_topic: str,
        start_time: str = None,
        end_time: str = None,
        duration: int = None
    ) -> bytes:
        """
        使用WeasyPrint将Markdown转换为高质量PDF
        
        Args:
            markdown_text: Markdown文本
            title: PDF标题
            debate_topic: 辩题
            start_time: 开始时间
            end_time: 结束时间
            duration: 持续时间（分钟）
            
        Returns:
            PDF字节流
        """
        try:
            meta_info = {
                "辩题": debate_topic,
            }
            if start_time:
                meta_info["开始时间"] = start_time
            if end_time:
                meta_info["结束时间"] = end_time
            if duration is not None:
                meta_info["持续时间"] = f"{duration} 分钟"

            pdf_bytes = await markdown_to_pdf(
                markdown_text=markdown_text,
                title=title,
                meta_info=meta_info,
                syntax_style="github"
            )
            return pdf_bytes
        except RuntimeError as e:
            logger.error(f"WeasyPrint 不可用，回退到 ReportLab: {e}", exc_info=True)
            return await asyncio.to_thread(
                ReportGenerator._markdown_to_pdf_bytes,
                markdown_text,
                title=title
            )
        except Exception as e:
            logger.error(f"WeasyPrint 转换PDF失败，回退到 ReportLab: {e}", exc_info=True)
            return await asyncio.to_thread(
                ReportGenerator._markdown_to_pdf_bytes,
                markdown_text,
                title=title
            )

    @staticmethod
    async def export_to_pdf_async(
        db: Session,
        debate_id: str,
        debate_topic: str,
        content_str: str = "",
        start_time: str = None,
        end_time: str = None,
        duration: int = None
    ) -> Optional[bytes]:
        """
        异步导出辩论报告为PDF（使用WeasyPrint）

        Args:
            db: 数据库会话
            debate_id: 辩论ID
            debate_topic: 辩题
            content_str: 辩论内容
            start_time: 开始时间
            end_time: 结束时间
            duration: 持续时间（分钟）

        Returns:
            PDF字节流
        """
        try:
            markdown_text = await ReportGenerator.generate_markdown_report_async(
                db=db,
                debate_topic=debate_topic,
                content_str=content_str,
            )
            if not markdown_text:
                return None

            return await ReportGenerator.render_markdown_to_pdf_async(
                markdown_text=markdown_text,
                debate_topic=debate_topic,
                start_time=start_time,
                end_time=end_time,
                duration=duration,
            )

        except Exception as e:
            logger.error(f"导出PDF失败: {e}", exc_info=True)
            return None

    @staticmethod
    async def generate_markdown_report_async(
        db: Session,
        debate_topic: str,
        content_str: str = "",
    ) -> Optional[str]:
        try:
            judge_prompt = f"""
你是智能辩论系统的裁判，请基于以下辩论记录生成一份 Markdown 复盘报告。

辩题：{debate_topic}
辩论详情：{content_str}

要求：
1. 使用中文输出。
2. 包含总体评价、双方核心论点、关键回合分析、逐个辩手优缺点和改进建议。
3. 如果记录中包含评分，请结合评分原因进行解释。
"""

            judgeagent = JudgeAgent(db)
            markdown_text = await judgeagent._call_agent(prompt=judge_prompt)

            try:
                data = json.loads(markdown_text)
                markdown_text = data.get("report")
            except Exception:
                logger.warning("报告 JSON 解析失败，使用原始文本")

            if not markdown_text:
                logger.error("生成的Markdown报告为空")
                return None

            return markdown_text
        except Exception as e:
            logger.error(f"生成Markdown报告失败: {e}", exc_info=True)
            return None

    @staticmethod
    async def render_markdown_to_pdf_async(
        markdown_text: str,
        debate_topic: str,
        start_time: str = None,
        end_time: str = None,
        duration: int = None,
    ) -> Optional[bytes]:
        try:
            return await ReportGenerator._markdown_to_pdf_bytes_weasyprint(
                markdown_text=markdown_text,
                title=f"辩论报告 - {debate_topic}",
                debate_topic=debate_topic,
                start_time=start_time,
                end_time=end_time,
                duration=duration,
            )
        except Exception as e:
            logger.error(f"渲染PDF失败: {e}", exc_info=True)
            return None

    
    @staticmethod
    def export_to_excel(report: Report) -> Optional[bytes]:
        """
        导出为Excel
        
        Args:
            report: 报告对象
            
        Returns:
            Excel文件字节流
        """
        try:
            workbook = Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = "报告概览"

            bold_font = Font(bold=True)
            title_font = Font(bold=True, size=14)

            summary_sheet["A1"] = "辩论报告"
            summary_sheet["A1"].font = title_font
            summary_rows = [
                ("辩题", report.topic),
                ("开始时间", report.start_time),
                ("结束时间", report.end_time),
                ("持续时间（分钟）", report.duration),
                ("获胜方", report.winner),
            ]
            for row_index, (label, value) in enumerate(summary_rows, start=3):
                summary_sheet.cell(row=row_index, column=1, value=label).font = bold_font
                summary_sheet.cell(row=row_index, column=2, value=value)

            participants_sheet = workbook.create_sheet("参与者")
            participants_sheet.append(["姓名", "角色", "立场", "最终得分"])
            for cell in participants_sheet[1]:
                cell.font = bold_font

            for participant in report.participants:
                final_score = participant.get("final_score") or {}
                participants_sheet.append([
                    participant.get("name", ""),
                    participant.get("role", ""),
                    participant.get("stance", ""),
                    final_score.get("overall_score", ""),
                ])

            speeches_sheet = workbook.create_sheet("发言记录")
            speeches_sheet.append(["序号", "环节", "内容", "时长", "得分"])
            for cell in speeches_sheet[1]:
                cell.font = bold_font

            for index, speech in enumerate(report.speeches, start=1):
                score = speech.get("score", {}).get("overall_score", "N/A") if speech.get("score") else "N/A"
                speeches_sheet.append([
                    index,
                    speech.get("phase", ""),
                    speech.get("content", ""),
                    speech.get("duration", ""),
                    score,
                ])

            statistics_sheet = workbook.create_sheet("统计")
            statistics_sheet.append(["指标", "值"])
            for cell in statistics_sheet[1]:
                cell.font = bold_font
            for key, value in (report.statistics or {}).items():
                statistics_sheet.append([
                    key,
                    json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value,
                ])

            for sheet in workbook.worksheets:
                for column_cells in sheet.columns:
                    column_letter = column_cells[0].column_letter
                    max_length = max(len(str(cell.value or "")) for cell in column_cells)
                    sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        
        except Exception as e:
            logger.error(f"导出Excel失败: {e}", exc_info=True)
            return None
